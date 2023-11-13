from openpyxl import Workbook
import openpyxl
import random
def f(file_path, s):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active
    ret = []
    for row_id, row in enumerate(ws.iter_rows()):
        for cell in row:
            if cell.value == s:
                ret.append(row_id + 1) #change column numb
                break
    return ret



workbook = openpyxl.Workbook()

# 选择默认的工作表
sheet = workbook.active
sheet.title = "Sheet1"
# 在单元格A1写入数据
l = ['s', 'ads', 'adsd', 'asdw', '123', '111', 'q', 'wer2', '13']
for x in range(1, 50):
    for y in ['A', 'B', 'C', 'D']:
        sheet[y + str(x)] = random.choice(l)

workbook.save('example.xlsx')
assert filterExcelRows('example.xlsx', 'ads') == f('example.xlsx', 'ads')